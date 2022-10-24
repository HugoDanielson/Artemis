import random

import openpyxl
import pickle
import pandas as pd
import subprocess
import os
import keyboard
import time



wb = openpyxl.load_workbook(("ExcelFiles/measureAutoImportTemplate.xlsx"))

ws = wb.active

OperatorOverride = False


def write_header(tipo, ws):
    ws['B1'] = tipo


def write_measurement(ws, tipo,df,i):
    offset = 7

    for y in range(0,df.size):
        rf = df[y]
        x = 1
        print(df.head())
        print(rf)


        if rf[4].get('nom')== 0:
            Mvar = 1
            Mminmax = 0
            Mavg = 0
        else:
            Mvar = 1
            Mminmax = 1
            Mavg = 0

        ws.cell(y+offset, x, tipo)
        ws.cell(y+offset, x + 1, 0)  # Standard
        ws.cell(y+offset, x + 2, "SHAREDOP")  # Extract from data
        ws.cell(y+offset, x + 3, 0)  # 1 if IR
        ws.cell(y+offset, x + 4, rf[9].get('M_num'))  # 10 if SPC 62
        ws.cell(y+offset, x + 5, rf[3].get('Measurement'))  # MeasurementName
        ws.cell(y+offset, x + 6, 0)  # measmode
        ws.cell(y+offset, x + 7, rf[4].get('nom')) # nominalMeasure
        ws.cell(y+offset, x + 8, rf[offset].get('min'))  # tol min
        ws.cell(y+offset, x + 9, rf[6].get('max'))  # tolmax
        ws.cell(y+offset, x + 10, rf[7].get('delta'))  # toldelta
        ws.cell(y+offset, x + 11, rf[4].get('nom'))  # Master
        ws.cell(y+offset, x + 12, 0)  # correction
        ws.cell(y+offset, x + 13, 0)  # calibrationmode
        ws.cell(y+offset, x + 14, 60)  # time/amount
        ws.cell(y+offset, x + 15,   -100)  # comparatormin
        ws.cell(y+offset, x + 16, 100)  # ComparatorMax
        ws.cell(y+offset, x + 17, 1)  # waitingtime
        ws.cell(y+offset, x + 18, 3)  # meastime
        ws.cell(y+offset, x + 19, 1)  # spc group
        ws.cell(y+offset, x + 20, "µm")  # unit of measure
        ws.cell(y+offset, x + 21, rf[9].get('M_num'))  # OpID
        ws.cell(y+offset, x + 22, 0)  # unilateral
        ws.cell(y+offset, x + 23, 0)  # fixedcontrollimit
        ws.cell(y+offset, x + 24, -100)  # lowerlimit
        ws.cell(y+offset, x + 25, 100)  # upperlimit
        ws.cell(y+offset, x + 26, "Beda{}".format(i))  # bedacode
        ws.cell(y+offset, x + 27, 1)  # step
        ws.cell(y+offset, x + 28, "2;3;8")  # notification
        ws.cell(y+offset, x + 29, 1)  # multipointcounter
        ws.cell(y+offset, x + 30, Mminmax)  # considermintol
        ws.cell(y+offset, x + 31, Mminmax)  # considermaxtol
        ws.cell(y+offset, x + 32, Mavg)  # consideraverage
        ws.cell(y+offset, x + 33, Mvar)  # considerVariation
        ws.cell(y+offset, x + 34, "empty")  # criticalchar
        ws.cell(y+offset, x + 35, 0)  # percentofbatch
        ws.cell(y+offset, x + 36, "AAAAA")  # ProcedureNumber


for i in range(1,3):
    bdata = "BearingData{}.p".format(i)
    bearing_data = pickle.load(open(bdata, "rb"))
    df = pd.DataFrame(bearing_data)

    for x in range(0,len(df.head().keys())):


        tipo = df.head().keys()[x]
        rf = df[tipo]



        #write_header(tipo, ws)
        #write_measurement(ws, tipo, rf, x)

        #name = "Tipo/{}.xlsx".format(tipo)
        #wb.save(name)









        if OperatorOverride == True:
            os.startfile("Tipo\{}.xlsx".format(tipo))
            time.sleep(2)
            keyboard.press('ctrl')
            keyboard.press('s')
            time.sleep(0.5)
            keyboard.release('ctrl')
            keyboard.release('s')
            time.sleep(0.2)
            keyboard.press('alt')
            keyboard.press('f')
            keyboard.press('x')
            time.sleep(0.5)
            keyboard.release('alt')
            keyboard.release('f')
            keyboard.release('x')




