import random

import openpyxl
import pickle
import pandas as pd
import subprocess
import os
import keyboard
import time

def run_script():


    wb = openpyxl.load_workbook(("ExcelFiles/measureAeroImportTemplate.xlsx"))

    ws = wb.active

    offset = 7

    for i in range(1,3):
        bdata = "BearingData{}.p".format(i)
        bearing_data = pickle.load(open(bdata, "rb"))
        df = pd.DataFrame(bearing_data)

        for x in range(0,len(df.head().keys())):

            tipo = df.head().keys()[x]
            rf = df[tipo]

            for y in range(0, rf.size):

                pf = rf[y]
                x = 1
                #print(df.head())
                #print(rf)

                if pf[4].get('nom') == 0:
                    Mvar = 1
                    Mminmax = 0
                    Mavg = 0
                elif pf[9].get('M_num') == 27 and pf[4].get('nom') != 0:
                    Mvar = 0
                    Mminmax = 1
                    Mavg = 0
                else:
                    Mvar = 1
                    Mminmax = 1
                    Mavg = 0

                if pf[5].get('min') < 0:
                    k = -1
                    j = 1
                elif pf[5].get('min') == 0:
                    k = 1
                    j = -1
                else:
                    k= 1
                    j = 1

                if pf[6].get('max') == 0:
                    max = pf[7].get('delta')
                else:
                    max = pf[6].get('max')

                ws.cell(y + offset, x, tipo) # Part Number
                ws.cell(y + offset, x + 2, pf[3].get('Measurement'))  # Measurement
                ws.cell(y + offset, x + 3, 0) #Codmac / Aero Facility Code in ERP
                ws.cell(y + offset, x + 4, pf[9].get('M_num'))  # 10 if SPC 62
                #ws.cell(y + offset, x + 5, 0) #OperationNumber/Routing
                ws.cell(y + offset, x + 6, pf[3].get('Measurement'))  # MeasurementName
                ws.cell(y + offset, x + 7, 0)  # Measmode
                ws.cell(y + offset, x + 8, 0)  # InspectionFrequency
                #ws.cell(y + offset, x + 9, "Beda{}".format(y)) #Dim ID
                ws.cell(y + offset, x + 10, "Beda{}".format(random.randint(1,9999)))  # bedacode
                ws.cell(y + offset, x + 11, pf[4].get('nom')) #Nominal Measure (Inch/mm)
                ws.cell(y + offset, x + 12, pf[5].get('min'))  # tol min
                ws.cell(y + offset, x + 13, max)  # tolmax
                ws.cell(y + offset, x + 14, pf[7].get('delta'))  # toldelta
                ws.cell(y + offset, x + 15, "µm")  # unit of measure
                ws.cell(y + offset, x + 16, 'empty')  # Ctype
                ws.cell(y + offset, x + 17, 1)  # calibrationmode
                ws.cell(y + offset, x + 18, 5)  # time/amount
                ws.cell(y + offset, x + 19, k * j * -150)  # comparatormin
                ws.cell(y + offset, x + 20, k* 150)  # ComparatorMax
                ws.cell(y + offset, x + 21, 1)  # waitingtime
                ws.cell(y + offset, x + 22, 8)  # meastime
                ws.cell(y + offset, x + 23, 1) #Group by (for spc)
                ws.cell(y + offset, x + 24, 1)  #Multipoint
                ws.cell(y + offset, x + 25, 0) #Unilateral
                ws.cell(y + offset, x + 26, 1) #Step (min. value 1)










            offset = offset + rf.size

            print(offset)
        name = "NewzFile.xlsx"
        wb.save(name)
            #write_header(tipo, ws)
            #write_measurement(ws, tipo, rf, x)